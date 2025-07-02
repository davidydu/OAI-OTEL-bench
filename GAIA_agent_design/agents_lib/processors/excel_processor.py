from __future__ import annotations

from typing import List

from ..file_router import Attachment

from openpyxl import load_workbook
from openpyxl.styles.colors import Color, COLOR_INDEX
from xml.etree import ElementTree as ET
import csv
import io


class ExcelProcessorAgent:
    """Convert Excel sheets to CSV-formatted text with style information."""

    def _tint_hex(self, hex_color: str, tint: float) -> str:
        """Apply Excel tint to a ``RRGGBB`` color string."""
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        if tint < 0:
            r = int(r * (1 + tint))
            g = int(g * (1 + tint))
            b = int(b * (1 + tint))
        else:
            r = int(r + (255 - r) * tint)
            g = int(g + (255 - g) * tint)
            b = int(b + (255 - b) * tint)
        return f"{r:02X}{g:02X}{b:02X}"

    
    def _tint_hex(self, hex_color: str, tint: float) -> str:
        """Apply Excel tint to a ``RRGGBB`` color string."""
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        if tint < 0:
            r = int(r * (1 + tint))
            g = int(g * (1 + tint))
            b = int(b * (1 + tint))
        else:
            r = int(r + (255 - r) * tint)
            g = int(g + (255 - g) * tint)
            b = int(b + (255 - b) * tint)
        return f"{r:02X}{g:02X}{b:02X}"

    def _parse_theme(self, theme_bytes: bytes | None) -> list[str]:
        """Extract theme colors from ``Workbook.loaded_theme``."""
        if not theme_bytes:
            return []
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        xml = ET.fromstring(theme_bytes)
        scheme = xml.find(".//a:clrScheme", ns)
        if scheme is None:
            return []
        colors: list[str] = []
        for child in scheme:
            val = "000000"
            srgb = child.find("a:srgbClr", ns)
            if srgb is not None and "val" in srgb.attrib:
                val = srgb.attrib["val"]
            else:
                sys = child.find("a:sysClr", ns)
                if sys is not None:
                    val = sys.attrib.get("lastClr", val)
            colors.append(val)
        return colors

    def _get_hex_color(self, color: Color | None, theme: list[str] | None) -> str:
        """Return a hex string for an openpyxl ``Color``."""
        if color is None:
            return ""
        ctype = getattr(color, "type", None)
        if ctype == "rgb":
            rgb = color.rgb
            if hasattr(rgb, "rgb"):
                rgb = rgb.rgb
            if isinstance(rgb, str):
                return "#" + rgb[-6:]
        elif ctype == "theme" and theme:
            idx = color.theme
            if idx is not None and 0 <= idx < len(theme):
                base = theme[idx][-6:]
                if getattr(color, "tint", 0):
                    base = self._tint_hex(base, color.tint)
                return "#" + base
        elif ctype == "indexed":
            idx = color.indexed
            if idx is not None and 0 <= idx < len(COLOR_INDEX):
                return "#" + COLOR_INDEX[idx][-6:]
        return ""


    def process(self, att: Attachment) -> str:
        """Convert an Excel file to CSV-formatted text."""
        if att.path.exists():
            wb = load_workbook(att.path, data_only=False)
        else:
            from io import BytesIO
            wb = load_workbook(filename=BytesIO(att.bytes), data_only=False)
        theme_colors = self._parse_theme(getattr(wb, "loaded_theme", None))
        texts: List[str] = []
        for sheet in wb.worksheets:
            output = io.StringIO()
            writer = csv.writer(output)
            for row in sheet.iter_rows():
                vals = []
                for cell in row:
                    if cell.data_type == "f":
                        text = f"FORMULA:{cell.value}"
                    else:
                        text = cell.value if cell.value is not None else ""
                    color = self._get_hex_color(cell.fill.start_color, theme_colors)
                    if not color:
                        color = self._get_hex_color(cell.font.color, theme_colors)
                    if color:
                        vals.append(f"{text} [COLOR={color}]")
                    else:
                        vals.append(text)
                writer.writerow(vals)
            texts.append(f"Sheet: {sheet.title}\n{output.getvalue()}")
        return "\n".join(texts)

